"""
Report Exporter Module - Exports tables to Excel and other formats with Biologic branding
"""

import pandas as pd
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as XLImage
from openpyxl.chart import BarChart, LineChart, Reference
import os
import matplotlib
matplotlib.use('Agg')  # Non-interactive backend
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.ticker import MaxNLocator
import matplotlib.font_manager as fm

class ReportExporter:
    """Handles exporting tables to various formats with Biologic branding"""
    
    def __init__(self):
        self.workbook = None
        # Biologic brand colors
        self.BIOLOGIC_PRIMARY = "1F4D4D"      # Dark teal
        self.BIOLOGIC_SECONDARY = "577A7A"    # Medium teal
        self.BIOLOGIC_LIGHT = "9AAFAF"        # Light teal
        self.BIOLOGIC_BACKGROUND = "E4EAEA"   # Background grey
        # Biologic brand color for charts
        self.BIOLOGIC_GOLD = '#D4AF37'  # Gold for bat call charts
        self.BIOLOGIC_TEAL = '#577A7A'  # Teal for line charts
    
    def _setup_montserrat_font(self):
        """Setup Montserrat font for matplotlib charts"""
        available_fonts = [f.name for f in fm.fontManager.ttflist]
        
        if 'Montserrat' in available_fonts:
            plt.rcParams['font.family'] = 'Montserrat'
            return True
        else:
            # Fallback to similar sans-serif fonts
            plt.rcParams['font.family'] = 'sans-serif'
            plt.rcParams['font.sans-serif'] = ['Montserrat', 'Arial', 'Helvetica', 'DejaVu Sans']
            return False
    
    def export_to_excel(self, tables: dict, audit_reports: dict, config: dict, 
                       processed_data: dict = None, master_data_df: pd.DataFrame = None) -> bytes:
        """
        Export all tables to a single Excel workbook
        
        Args:
            tables: Dictionary of table data
            audit_reports: Dictionary of audit reports
            config: Configuration used to generate tables
            processed_data: Dictionary of processed site data for charts (optional)
            master_data_df: Master data DataFrame for bat activity statements (optional)
            
        Returns:
            BytesIO object containing Excel file
        """
        # Create new workbook
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)  # Remove default sheet
        
        # Add tables in order
        table_order = ['2.1', '3.2', '3.3', '3.4', '3.5', '3.6', '3.7', '3.8', 
                      '3.9', '3.10', '3.11', '3.12', '3.13', '3.14']
        
        for table_id in table_order:
            if table_id in tables:
                table_data = tables[table_id]
                self._add_table_sheet(table_data)
        
        # Add bat activity statements if master data provided
        if master_data_df is not None:
            self._add_bat_activity_statements(master_data_df)
        
        # Add audit sheets
        self._add_audit_sheets(audit_reports)
        
        # Add PLNB activity charts if processed data is provided
        if processed_data and config:
            print("\n=== STARTING CHART CREATION ===")
            print(f"processed_data provided: {processed_data is not None}")
            print(f"config provided: {config is not None}")
            try:
                # Add Figures 3.1-3.4: Microclimate charts for continuous sites
                microclimate_charts = self._add_microclimate_charts(processed_data, config)
                print(f"✅ Successfully added {microclimate_charts} microclimate charts (Figures 3.1-3.4)")
                
                # Add NEW: Temperature and Humidity charts (8 total: 4 temp + 4 humidity)
                temp_humidity_charts = self._add_temperature_humidity_charts(processed_data, config)
                print(f"✅ Successfully added {temp_humidity_charts} temperature and humidity charts")
                
                # Add Figure 3.5: Average annual PLNB activity
                annual_chart = self._add_annual_plnb_chart(processed_data, config)
                if annual_chart:
                    print(f"✅ Successfully added Figure 3.5: Average annual PLNB activity")
                
                # Add Figures 3.8-3.11: PLNB calls per night for continuous sites
                plnb_charts = self._add_plnb_charts(processed_data, config)
                if plnb_charts > 0:
                    print(f"✅ Successfully added {plnb_charts} PLNB charts (Figures 3.8-3.11)")
                else:
                    print("⚠️ No PLNB charts were created")
            except Exception as e:
                print(f"❌ ERROR: Could not add charts: {e}")
                import traceback
                traceback.print_exc()
        else:
            print(f"⚠️ Skipping charts - processed_data: {processed_data is not None}, config: {config is not None}")
        
        # Save to BytesIO
        excel_file = BytesIO()
        self.workbook.save(excel_file)
        excel_file.seek(0)
        
        return excel_file.getvalue()
    
    def _add_plnb_charts(self, processed_data: dict, config: dict) -> int:
        """
        Add four PLNB activity charts for continuous monitoring sites on a single sheet
        
        Args:
            processed_data: Dictionary with processed site data
            config: Configuration with site selections and date ranges
            
        Returns:
            Number of charts successfully created
        """
        # Get continuous monitoring sites from config
        continuous_sites = config.get('continuous_sites', [])
        
        if not continuous_sites:
            print("No continuous sites found in config")
            return 0
        
        print(f"Creating charts for sites: {continuous_sites}")
        print(f"Available keys in processed_data: {list(processed_data.keys())}")
        
        # Create a single worksheet for all charts
        ws = self.workbook.create_sheet(title="Figs3.8-3.11_PLNB")
        
        # Add title
        ws['A1'] = "Figures 3.8-3.11: Pilbara leaf-nosed bat calls per night (continuous monitoring sites)"
        ws['A1'].font = Font(bold=True, size=12)
        ws.merge_cells('A1:G1')
        
        # Process each site (up to 4)
        chart_count = 0
        chart_positions = ['A3', 'A20', 'A37', 'A54']  # Vertical stacking with spacing
        
        for idx, site_name in enumerate(continuous_sites[:4]):
            # Look for key with _continuous suffix
            site_key = f"{site_name}_continuous"
            
            print(f"Looking for site: {site_name} with key: {site_key}")
            
            if site_key in processed_data:
                site_data = processed_data[site_key]
                raw_df = site_data.get('raw_data')
                
                if raw_df is not None and not raw_df.empty:
                    start_date = site_data.get('start_date')
                    end_date = site_data.get('end_date')
                    chart_count += 1
                    print(f"✓ Creating chart {chart_count} for {site_name}")
                    try:
                        # Create chart and get image
                        img_data = self._create_plnb_chart_png(raw_df, site_name, chart_count, start_date, end_date)
                        
                        if img_data:
                            # Insert image into worksheet
                            img = XLImage(BytesIO(img_data))
                            img.width = 600  # Width in pixels
                            img.height = 350  # Height in pixels
                            ws.add_image(img, chart_positions[idx])
                            print(f"✅ Chart {chart_count} created successfully for {site_name}")
                        else:
                            print(f"❌ Failed to create chart image for {site_name}")
                            chart_count -= 1
                    except Exception as e:
                        print(f"❌ Error creating chart for {site_name}: {e}")
                        import traceback
                        traceback.print_exc()
                        chart_count -= 1
                else:
                    print(f"✗ No data available for {site_name}")
            else:
                print(f"✗ Site key '{site_key}' not found in processed data")
        
        print(f"Total charts created: {chart_count}")
        return chart_count
    
    def _create_plnb_chart_png(self, df: pd.DataFrame, site_name: str, chart_number: int, 
                          start_date=None, end_date=None):
        """
        Create a single PLNB activity smooth line chart with proper axis labels
        
        Args:
            df: DataFrame with raw site data
            site_name: Name of the site
            chart_number: Chart number (1-4)
            start_date: Start date for title (optional)
            end_date: End date for title (optional)
            
        Returns:
            bytes: PNG image data, or None if failed
        """
        try:
            # Find date column
            date_col = None
            for col in df.columns:
                if 'date' in str(col).lower():
                    date_col = col
                    break
            
            if date_col is None:
                date_col = df.columns[0]
            
            # Get column H (plnb_total_calls) - 8th column (index 7)
            plnb_col = None
            if len(df.columns) > 7:
                plnb_col = df.columns[7]  # Column H is index 7
            else:
                # Fallback: look for plnb_total_calls column by name
                for col in df.columns:
                    col_str = str(col).lower()
                    if 'plnb' in col_str and 'total' in col_str and 'call' in col_str:
                        plnb_col = col
                        break
            
            if plnb_col is None:
                print(f"Could not find PLNB calls column for {site_name}")
                return None
            
            print(f"Using date column: {date_col}, PLNB column: {plnb_col}")
            
            # Prepare data for matplotlib chart
            chart_data = []
            for idx, row in df.iterrows():
                date_val = row[date_col]
                calls_val = row[plnb_col]
                
                # Skip rows with no date
                if pd.isna(date_val):
                    continue
                
                # Convert date to datetime if not already
                if not isinstance(date_val, (datetime, pd.Timestamp)):
                    try:
                        date_val = pd.to_datetime(date_val)
                    except:
                        continue
                
                # Handle NaN or empty call values - treat as 0
                if pd.isna(calls_val) or calls_val == '':
                    calls_val = 0
                else:
                    try:
                        calls_val = float(calls_val)
                    except:
                        calls_val = 0
                
                chart_data.append({'date': date_val, 'calls': calls_val})
            
            if len(chart_data) == 0:
                print(f"No valid data rows for {site_name}")
                return None
            
            print(f"Creating matplotlib smooth line chart with {len(chart_data)} data points for {site_name}")
            
            # Convert to DataFrame for easier plotting
            chart_df = pd.DataFrame(chart_data)
            chart_df['date'] = pd.to_datetime(chart_df['date'])
            chart_df = chart_df.sort_values('date')
            
            # Create matplotlib smooth line chart with filled area
            self._setup_montserrat_font()
            
            fig, ax = plt.subplots(figsize=(10, 6))
            fig.patch.set_facecolor('white')
            ax.patch.set_facecolor('white')
            
            # Create smooth line with filled area using Biologic gold color
            # Fill the area under the curve with same solid color as line
            ax.fill_between(chart_df['date'], 0, chart_df['calls'], 
                           color=self.BIOLOGIC_GOLD, 
                           alpha=1.0,  # Solid color, same as line
                           zorder=1)
            
            # Then plot the smooth line on top
            ax.plot(chart_df['date'], chart_df['calls'], 
                   color=self.BIOLOGIC_GOLD,
                   linewidth=2.5,
                   zorder=2)
            
            # ========================================================================
            # FIX 1: Format X-axis to show mmm-yy
            # ========================================================================
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%b-%y'))
            ax.xaxis.set_major_locator(mdates.MonthLocator())
            plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha='right')
            
            # ========================================================================
            # FIX 2: Format Y-axis to show whole numbers only
            # ========================================================================
            ax.yaxis.set_major_locator(MaxNLocator(integer=True))
            
            # ========================================================================
            # FIX 3: Add axis labels with Montserrat font
            # ========================================================================
            ax.set_xlabel('Month', fontsize=11, fontweight='normal')
            ax.set_ylabel('Number of calls per night', fontsize=11, fontweight='normal')
            
            # Set chart title (site name)
            ax.set_title(f'{site_name}', fontsize=12, fontweight='bold', pad=15)
            
            # ========================================================================
            # Professional styling
            # ========================================================================
            # Remove top and right spines
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)
            
            # Make remaining spines thinner
            ax.spines['left'].set_linewidth(0.5)
            ax.spines['bottom'].set_linewidth(0.5)
            
            # Set tick parameters
            ax.tick_params(axis='both', which='major', labelsize=10, length=4, width=0.5)
            
            # Add subtle grid for y-axis only
            ax.yaxis.grid(True, linestyle='--', alpha=0.3, linewidth=0.5)
            ax.set_axisbelow(True)  # Put grid behind lines
            
            # Ensure y-axis starts at 0
            ax.set_ylim(bottom=0)
            
            # Adjust layout to prevent label cutoff
            plt.tight_layout()
            
            # Save figure to BytesIO
            buf = BytesIO()
            plt.savefig(buf, format='png', dpi=300, bbox_inches='tight', facecolor='white')
            plt.close(fig)
            buf.seek(0)
            
            print(f"✅ Successfully created matplotlib smooth line chart for {site_name}")
            return buf.getvalue()
            
        except Exception as e:
            print(f"❌ Error creating chart for {site_name}: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def _add_microclimate_charts(self, processed_data: dict, config: dict) -> int:
        """
        Add Figures 3.1-3.4: Microclimate comparison charts as high-resolution transparent PNGs
        
        Args:
            processed_data: Dictionary with processed site data
            config: Configuration with site selections
            
        Returns:
            Number of microclimate charts created
        """
        import matplotlib
        matplotlib.use('Agg')  # Non-interactive backend
        import matplotlib.pyplot as plt
        import matplotlib.dates as mdates
        from matplotlib.ticker import MultipleLocator
        from io import BytesIO
        
        # Configure matplotlib for high-resolution output
        plt.rcParams['figure.dpi'] = 600  # ✅ FIX #2: Doubled to 600 DPI
        plt.rcParams['savefig.dpi'] = 600
        plt.rcParams['font.family'] = 'sans-serif'
        plt.rcParams['font.sans-serif'] = ['Montserrat', 'DejaVu Sans', 'Arial']
        plt.rcParams['font.size'] = 10
        plt.rcParams['axes.linewidth'] = 0.8
        plt.rcParams['grid.linewidth'] = 0.5
        plt.rcParams['grid.alpha'] = 0.3
        
        continuous_sites = config.get('continuous_sites', [])
        if not continuous_sites:
            return 0
        
        chart_count = 0
        for idx, site_name in enumerate(continuous_sites[:4], 1):
            site_key = f"{site_name}_continuous"
            
            if site_key not in processed_data:
                continue
                
            site_data = processed_data[site_key]
            df = site_data.get('raw_data')
            
            if df is None or df.empty:
                continue
            
            print(f"✓ Creating high-res microclimate chart {idx} for {site_name}")
            
            # Create worksheet for data (optional - can be hidden)
            sheet_name = f"Fig3.{idx}_{site_name}"[:31]
            ws = self.workbook.create_sheet(title=sheet_name)
            
            # Generate the high-resolution PNG
            png_data = self._create_microclimate_png(df, site_name, site_data, idx)
            
            if png_data:
                # Add PNG image to worksheet
                from openpyxl.drawing.image import Image as XLImage
                img = XLImage(BytesIO(png_data))
                
                # Scale to fit nicely in Excel (convert cm to pixels at 96 DPI for Excel display)
                # 24cm × 16cm at 96 DPI ≈ 906 × 604 pixels for Excel
                img.width = 906
                img.height = 604
                
                ws.add_image(img, 'A1')
                chart_count += 1
                print(f"  ✅ Successfully created Figure 3.{idx}: {site_name}")
            else:
                print(f"  ❌ Failed to create chart for {site_name}")
        
        return chart_count
    
    def _create_microclimate_png(self, df, site_name, site_data, fig_number):
        """
        Create a high-resolution transparent PNG of microclimate data
        
        Returns:
            bytes: PNG image data, or None if failed
        """
        import matplotlib.pyplot as plt
        import matplotlib.dates as mdates
        from matplotlib.ticker import MultipleLocator
        import pandas as pd
        import numpy as np
        from io import BytesIO
        
        # Try to import scipy for smooth interpolation, fall back to simple methods if not available
        try:
            from scipy import interpolate
            from scipy.ndimage import gaussian_filter1d
            HAS_SCIPY = True
            HAS_GAUSSIAN = True
        except ImportError:
            HAS_SCIPY = False
            HAS_GAUSSIAN = False
            print("  ⚠️ scipy not available, will use simple interpolation for rainfall")
        
        try:
            # Prepare data
            date_col = df.columns[0]
            dates = pd.to_datetime(df[date_col], errors='coerce')
            
            # Calculate metrics (same as before)
            # Roost temps (columns V-AC, indices 21-28)
            roost_temp_cols = [21, 22, 23, 24, 25, 26, 27, 28]
            roost_temps = [pd.to_numeric(df.iloc[:, col], errors='coerce') 
                          for col in roost_temp_cols if col < len(df.columns)]
            mean_roost_temp = pd.concat(roost_temps, axis=1).mean(axis=1) if roost_temps else pd.Series()
            
            # Roost RH (columns AD-AK, indices 29-36)
            roost_rh_cols = [29, 30, 31, 32, 33, 34, 35, 36]
            roost_rhs = [pd.to_numeric(df.iloc[:, col], errors='coerce')
                        for col in roost_rh_cols if col < len(df.columns)]
            mean_roost_rh = pd.concat(roost_rhs, axis=1).mean(axis=1) if roost_rhs else pd.Series()
            
            # Rainfall (column S, index 18)
            rainfall = pd.to_numeric(df.iloc[:, 18], errors='coerce') if 18 < len(df.columns) else pd.Series()
            
            # Ambient temp (columns Q-R, indices 16-17)
            ambient_temps = [pd.to_numeric(df.iloc[:, col], errors='coerce')
                            for col in [16, 17] if col < len(df.columns)]
            ambient_temp = pd.concat(ambient_temps, axis=1).mean(axis=1) if ambient_temps else pd.Series()
            
            # Ambient RH (columns T-U, indices 19-20)
            ambient_rhs = [pd.to_numeric(df.iloc[:, col], errors='coerce')
                          for col in [19, 20] if col < len(df.columns)]
            ambient_rh = pd.concat(ambient_rhs, axis=1).mean(axis=1) if ambient_rhs else pd.Series()
            
            # Create figure with size 24cm × 16cm at 300 DPI
            # Convert cm to inches: 24cm = 9.45", 16cm = 6.3"
            fig, ax1 = plt.subplots(figsize=(9.45, 6.3))
            fig.patch.set_alpha(0)  # Transparent background
            ax1.patch.set_alpha(0)
            
            # NO TITLE (removed as requested)
            
            # Create second y-axis for rainfall
            ax2 = ax1.twinx()
            
            # Plot rainfall - try monthly smooth trendline, fallback to bars if needed
            if not rainfall.empty and rainfall.notna().any():
                try:
                    # Create DataFrame for resampling
                    rainfall_df = pd.DataFrame({'date': dates, 'rainfall': rainfall})
                    rainfall_df = rainfall_df.dropna(subset=['date'])
                    rainfall_df['date'] = pd.to_datetime(rainfall_df['date'])
                    rainfall_df = rainfall_df.set_index('date')
                    
                    # Resample to monthly totals (sum of rainfall per month)
                    monthly_rainfall = rainfall_df['rainfall'].resample('M').sum()
                    
                    if not monthly_rainfall.empty and monthly_rainfall.sum() > 0 and len(monthly_rainfall) > 1:
                        # Get monthly dates (end of month)
                        monthly_dates = monthly_rainfall.index
                        monthly_values = monthly_rainfall.values
                        
                        # Convert dates to numeric for interpolation
                        date_nums = mdates.date2num(monthly_dates)
                        
                        # Create smooth interpolation
                        if HAS_SCIPY and len(date_nums) > 3:
                            # Use cubic spline for smooth curve (requires scipy)
                            # Changed to 'clamped' boundary for smoother edges
                            cs = interpolate.CubicSpline(date_nums, monthly_values, bc_type='clamped')
                            
                            # Generate MORE points for SMOOTHER curve (increased from 300 to 800)
                            date_nums_smooth = np.linspace(date_nums.min(), date_nums.max(), 800)
                            rainfall_smooth = cs(date_nums_smooth)
                            # Ensure no negative values from interpolation
                            rainfall_smooth = np.maximum(rainfall_smooth, 0)
                            
                            # Apply light Gaussian smoothing for extra smoothness (if available)
                            if HAS_GAUSSIAN:
                                rainfall_smooth = gaussian_filter1d(rainfall_smooth, sigma=3)
                                rainfall_smooth = np.maximum(rainfall_smooth, 0)  # Re-check after smoothing
                            
                            dates_smooth = mdates.num2date(date_nums_smooth)
                        else:
                            # Use linear interpolation (no scipy needed) with MANY more points for smoothness
                            # Increase from 300 to 800 points to match scipy version
                            date_nums_smooth = np.linspace(date_nums.min(), date_nums.max(), 800)
                            rainfall_smooth = np.interp(date_nums_smooth, date_nums, monthly_values)
                            
                            # Apply simple moving average for extra smoothness (scipy-free alternative)
                            # This approximates Gaussian smoothing
                            window_size = 15  # Smoothing window
                            rainfall_smooth = np.convolve(rainfall_smooth, 
                                                         np.ones(window_size)/window_size, 
                                                         mode='same')
                            # Ensure no negative values after smoothing
                            rainfall_smooth = np.maximum(rainfall_smooth, 0)
                            
                            dates_smooth = mdates.num2date(date_nums_smooth)
                        
                        # Plot smooth line with MORE translucent fill (behind ALL other series)
                        # Blue color #4472C4 with HIGH transparency
                        # z-order -1 to GUARANTEE it's behind everything including gridlines
                        ax2.plot(dates_smooth, rainfall_smooth, color='#4472C4', linewidth=1.5, 
                                label='Rainfall (mm)', zorder=-1, alpha=0.5)  # Reduced line opacity
                        ax2.fill_between(dates_smooth, 0, rainfall_smooth, 
                                        color='#4472C4', alpha=0.15, zorder=-1)  # 50% more transparent (was 0.3)
                    else:
                        # Not enough data for smooth curve, use simple bar chart
                        raise ValueError("Insufficient data for smooth interpolation")
                        
                except Exception as e:
                    # Fallback to simple bar chart if smoothing fails
                    print(f"  ⚠️ Rainfall smoothing failed ({e}), using bar chart fallback")
                    rainfall_clean = rainfall.fillna(0)
                    # Simple bars with consistent blue color (matching transparency)
                    ax2.bar(dates, rainfall_clean, width=1.0, color='#4472C4', 
                           alpha=0.15, edgecolor='none', label='Rainfall (mm)', zorder=-1)
            
            # Plot temperature lines (solid, thicker)
            # SWAPPED COLORS: Ambient=Red, Roost=Orange
            if not ambient_temp.empty:
                ax1.plot(dates, ambient_temp, color='#E74C3C', linewidth=2.0, 
                        label='Ambient temperature (°C)', zorder=3)  # Red
            
            if not mean_roost_temp.empty:
                ax1.plot(dates, mean_roost_temp, color='#D97E2A', linewidth=2.0,
                        label='Roost temperature (°C)', zorder=3)  # Orange
            
            # Plot humidity lines (dashed)
            if not ambient_rh.empty:
                ax1.plot(dates, ambient_rh, color='#70AD47', linewidth=1.5, 
                        linestyle='--', label='Ambient RH (%)', zorder=2)
            
            if not mean_roost_rh.empty:
                ax1.plot(dates, mean_roost_rh, color='#FFC000', linewidth=1.5,
                        linestyle='--', label='Roost RH (%)', zorder=2)
            
            # Configure left Y-axis (Temperature and Humidity)
            # Increased to 120 to match rainfall scale and accommodate high humidity
            ax1.set_ylabel('Temperature (°C) and Relative Humidity (%)', fontsize=10)
            ax1.set_ylim(0, 120)
            ax1.yaxis.set_major_locator(MultipleLocator(20))  # Tick marks every 20
            ax1.grid(True, alpha=0.3, linestyle='-', linewidth=0.5)
            ax1.tick_params(axis='y', labelsize=9)
            
            # Configure right Y-axis (Rainfall)
            # Increased to 120mm to capture rainfall events over 100mm
            ax2.set_ylabel('Rainfall (mm)', fontsize=10)
            ax2.set_ylim(0, 120)
            ax2.yaxis.set_major_locator(MultipleLocator(20))  # Tick marks every 20 to match left axis
            ax2.tick_params(axis='y', labelsize=9)
            
            # CRITICAL: Bring ax1 (temperature/humidity) to front so it draws over ax2 (rainfall)
            ax1.set_zorder(ax2.get_zorder() + 1)  # ax1 on top
            ax1.patch.set_visible(False)  # Make ax1 background transparent so ax2 shows through
            
            # Configure X-axis to capture full extent of data
            ax1.set_xlabel('')  # No x-axis label
            ax1.xaxis.set_major_formatter(mdates.DateFormatter('%b-%y'))
            ax1.xaxis.set_major_locator(mdates.MonthLocator())
            ax1.tick_params(axis='x', labelsize=9, rotation=0)
            
            # Set X-axis limits to full data extent
            if not dates.empty:
                date_min = dates.min()
                date_max = dates.max()
                if pd.notna(date_min) and pd.notna(date_max):
                    ax1.set_xlim(date_min, date_max)
            
            # Create legend with specific order:
            # [Ambient temp] [Ambient RH] | Rainfall | [Roost temp] [Roost RH]
            handles1, labels1 = ax1.get_legend_handles_labels()
            handles2, labels2 = ax2.get_legend_handles_labels()
            
            # Reorder: Ambient conditions, then Rainfall, then Roost conditions
            ordered_handles = []
            ordered_labels = []
            
            # Add Ambient temperature first
            for h, l in zip(handles1, labels1):
                if 'Ambient temperature' in l:
                    ordered_handles.append(h)
                    ordered_labels.append(l)
            
            # Add Ambient RH second
            for h, l in zip(handles1, labels1):
                if 'Ambient RH' in l:
                    ordered_handles.append(h)
                    ordered_labels.append(l)
            
            # Add Rainfall third (middle position)
            for h, l in zip(handles2, labels2):
                if 'Rainfall' in l:
                    ordered_handles.append(h)
                    ordered_labels.append(l)
            
            # Add Roost temperature fourth
            for h, l in zip(handles1, labels1):
                if 'Roost temperature' in l:
                    ordered_handles.append(h)
                    ordered_labels.append(l)
            
            # Add Roost RH fifth
            for h, l in zip(handles1, labels1):
                if 'Roost RH' in l:
                    ordered_handles.append(h)
                    ordered_labels.append(l)
            
            # ✅ FIX #1: Replace rainfall handle with custom blue rectangle patch
            from matplotlib.patches import Rectangle
            for i, (h, l) in enumerate(zip(ordered_handles, ordered_labels)):
                if 'Rainfall' in l:
                    ordered_handles[i] = Rectangle((0, 0), 1, 1, fc='#4472C4', 
                                                   alpha=0.7, edgecolor='none')
            
            # Place legend in ONE LINE at bottom center
            ax1.legend(ordered_handles, ordered_labels, 
                      loc='upper center', bbox_to_anchor=(0.5, -0.08),
                      ncol=5,  # All 5 items in one row
                      frameon=False, fontsize=9)
            
            # Tight layout with room for legend at bottom
            plt.tight_layout()
            plt.subplots_adjust(bottom=0.15)
            
            # Save to bytes
            buf = BytesIO()
            plt.savefig(buf, format='png', dpi=600, transparent=True, bbox_inches='tight')  # ✅ FIX #2: 600 DPI
            plt.close(fig)
            buf.seek(0)
            
            return buf.getvalue()
            
        except Exception as e:
            print(f"  ❌ Error creating PNG for {site_name}: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def _add_annual_plnb_chart(self, processed_data: dict, config: dict) -> bool:
        series.graphicalProperties = GraphicalProperties()
        series.graphicalProperties.line = LineProperties(w=25400)  # Line width
        series.graphicalProperties.line.solidFill = ColorChoice(srgbClr=color_hex)
        series.graphicalProperties.shadow = None
        
        # Set series title using SeriesLabel
        from openpyxl.chart.series import SeriesLabel
        series.tx = SeriesLabel(strRef=None)
        series.tx.v = title
    
    def _add_annual_plnb_chart(self, processed_data: dict, config: dict) -> bool:
        """
        Add Figure 3.5: PLNB calls per recording night with logarithmic Y-axis
        Creates charts for ALL SSM sites in a dynamic grid layout
        """
        from openpyxl.chart import ScatterChart, Reference, Series
        from openpyxl.chart.shapes import GraphicalProperties
        from openpyxl.drawing.line import LineProperties
        from openpyxl.drawing.fill import ColorChoice
        from openpyxl.chart.axis import Scaling
        
        # Get SSM sites from config - handle both list and dict formats
        ssm_config = config.get('ssm_sites', [])
        
        if isinstance(ssm_config, dict):
            # Config has dictionary format: {'CMPC-03': {'start': ..., 'end': ...}, ...}
            ssm_site_names = list(ssm_config.keys())
            print(f"Found {len(ssm_site_names)} SSM sites from config (dict format): {ssm_site_names}")
        elif isinstance(ssm_config, list):
            # Config has list format: ['CMPC-03', 'CMPC-08', ...]
            ssm_site_names = ssm_config
            print(f"Found {len(ssm_site_names)} SSM sites from config (list format): {ssm_site_names}")
        else:
            ssm_site_names = []
            print(f"⚠️ SSM sites config is neither dict nor list: {type(ssm_config)}")
        
        if not ssm_site_names:
            # Fallback: find sites with _ssm suffix in processed_data
            ssm_sites = [site for site in processed_data.keys() if site.endswith('_ssm')]
            print(f"⚠️ No SSM sites in config, found {len(ssm_sites)} in processed_data")
        else:
            # Build site keys from config names (add _ssm suffix)
            ssm_sites = [f"{name}_ssm" for name in ssm_site_names]
            print(f"Processing ALL {len(ssm_sites)} SSM sites: {ssm_sites}")
        
        if not ssm_sites:
            print("⚠️ No SSM sites available")
            return False
        
        ws = self.workbook.create_sheet(title="Fig3.5_PLNB_LogScale")
        ws['A1'] = "Figure 3.5: Pilbara leaf-nosed bat calls per recording night (logarithmic scale)"
        ws['A1'].font = Font(bold=True, size=11)
        ws.merge_cells('A1:Z1')
        
        chart_count = 0
        current_col = 1
        
        # Process ALL sites (removed [:4] limit)
        sites_found = []
        sites_missing = []
        sites_no_data = []
        
        for idx, site_key in enumerate(ssm_sites, 1):
            # Extract site name (remove _ssm suffix for display)
            site_name = site_key.replace('_ssm', '')
            
            print(f"  → Processing SSM site {idx}/{len(ssm_sites)}: {site_name} (key: {site_key})")
            
            if site_key not in processed_data:
                print(f"     ✗ Site key '{site_key}' not found in processed_data")
                sites_missing.append(site_name)
                continue
            
            site_data = processed_data[site_key]
            df = site_data.get('raw_data')
            
            # IMPORTANT: Create chart even if no data - shows monitoring occurred
            has_data = df is not None and not df.empty
            
            if not has_data:
                print(f"     ⚠️ No data for {site_name} - creating empty chart to show monitoring occurred")
                sites_no_data.append(site_name)
                # Create empty dataframe with SSM date range from config
                if isinstance(ssm_config, dict) and site_name in ssm_config:
                    start_date = pd.to_datetime(ssm_config[site_name]['start'])
                    end_date = pd.to_datetime(ssm_config[site_name]['end'])
                    # Create a date range for the monitoring period
                    dates = pd.date_range(start=start_date, end=end_date, freq='D')
                    df = pd.DataFrame({
                        'date': dates,
                        'plnb_total_calls': [0] * len(dates)  # All zeros - no calls detected
                    })
                    # Add dummy columns to match expected structure
                    for i in range(10):
                        if i not in [0, 7]:  # Skip date (0) and plnb (7) columns
                            df.insert(i, f'col_{i}', None)
                else:
                    print(f"     ✗ Cannot create empty chart - no date range in config for {site_name}")
                    sites_missing.append(site_name)
                    continue
            
            # Find date column
            date_col = None
            for col in df.columns:
                if 'date' in str(col).lower():
                    date_col = col
                    break
            if date_col is None:
                date_col = df.columns[0]
            
            # Find PLNB column (column 7 or by name)
            plnb_col = None
            if len(df.columns) > 7:
                plnb_col = df.columns[7]
            else:
                # Look for plnb_total_calls by name
                for col in df.columns:
                    if 'plnb' in str(col).lower() and 'total' in str(col).lower():
                        plnb_col = col
                        break
            
            if plnb_col is None:
                print(f"     ✗ No PLNB column found for {site_name}")
                sites_missing.append(site_name)
                continue
            
            # Filter and prepare data
            df_filtered = df[[date_col, plnb_col]].copy()
            df_filtered[date_col] = pd.to_datetime(df_filtered[date_col], errors='coerce')
            df_filtered = df_filtered[df_filtered[date_col].notna()]
            
            # Handle empty or all-null call data
            if df_filtered.empty or df_filtered[plnb_col].isna().all():
                print(f"     ⚠️ All data is null for {site_name} - using zeros")
                # Keep the dates but set calls to 0
                df_filtered[plnb_col] = 0
            
            # Write headers
            ws.cell(row=3, column=current_col, value='Recording night').font = Font(bold=True)
            ws.cell(row=3, column=current_col + 1, value='Number of (log) calls').font = Font(bold=True)
            
            # Write data rows
            row_num = 4
            # ✅ FIX #4: Track data characteristics for proper scaling
            has_positive = False  # Has any value > 0
            has_zero = False      # Has any value == 0
            all_zero = True       # All values are 0 (no positive values)
            
            for _, row in df_filtered.iterrows():
                date_val = row[date_col]
                calls_val = float(row[plnb_col]) if pd.notna(row[plnb_col]) else 0
                
                # Track data characteristics for axis scaling
                if calls_val > 0:
                    has_positive = True
                    all_zero = False
                elif calls_val == 0:
                    has_zero = True
                
                # ✅ FIX #4 CRITICAL: Keep actual value, don't convert 0 to 1!
                # OLD BUG: max(1, calls_val) converted all zeros to ones
                actual_calls_val = calls_val  # Keep zeros as zeros
                
                date_cell = ws.cell(row=row_num, column=current_col, value=date_val)
                calls_cell = ws.cell(row=row_num, column=current_col + 1, value=actual_calls_val)
                date_cell.number_format = 'dd-mmm'
                row_num += 1
            
            max_row = row_num - 1
            
            # Create chart even if only 1 data point
            if max_row < 4:
                print(f"     ⚠️ No data rows written for {site_name}, skipping chart")
                sites_missing.append(site_name)
                continue
            
            # ✅ FIX #4: Determine chart type based on actual data
            is_empty = not has_positive
            
            # Special sites that get watermark when empty
            watermark_sites = ['CMPC-13', 'CMPC-21', 'WMPC-21', 'WMPC-35']
            needs_watermark = is_empty and site_name in watermark_sites
            
            print(f"     Data analysis: has_positive={has_positive}, has_zero={has_zero}, all_zero={all_zero}")
            if needs_watermark:
                print(f"     → Will add watermark for {site_name}")
            
            chart = ScatterChart()
            
            # Chart title: Site name in Montserrat SemiBold, 12pt
            chart.title = site_name
            from openpyxl.chart.text import RichText
            from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font as DrawingFont
            
            # Format title with Montserrat SemiBold 12pt
            if chart.title:
                title_text = RichText()
                para = Paragraph()
                para.pPr = ParagraphProperties()
                run_props = CharacterProperties()
                run_props.sz = 1200  # 12pt in 1/100th of a point
                run_props.latin = DrawingFont(typeface='Montserrat SemiBold')
                para.pPr.defRPr = run_props
                title_text.p = [para]
                # Note: openpyxl has limited rich text support, title styling may need manual adjustment
            
            chart.style = 2
            
            # Y-axis title: "Number of (log) calls" - Montserrat 9pt
            chart.y_axis.title = "Number of (log) calls"
            
            # X-axis title: "Recording night" - Montserrat 9pt
            chart.x_axis.title = "Recording night"
            
            # ✅ FIX #4: Proper axis scaling based on ACTUAL data characteristics
            # This matches the behavior table from the Quick Guide
            if all_zero:
                # All zeros: linear 0-1 scale (shows flat line at 0)
                chart.y_axis.scaling = Scaling(logBase=None, min=0, max=1, orientation="minMax")
                print(f"     Axis: Linear scale 0-1 (all zeros - flat line at bottom)")
            elif has_zero:
                # Has zeros mixed with positive values: linear scale 0-auto
                chart.y_axis.scaling = Scaling(logBase=None, min=0, orientation="minMax")
                print(f"     Axis: Linear scale 0-auto (has zeros + positives)")
            elif has_positive and not has_zero:
                # All positive (no zeros): logarithmic scale 1-100
                chart.y_axis.scaling = Scaling(logBase=10, min=1, max=100, orientation="minMax")
                print(f"     Axis: Logarithmic scale 1-100 (all positive, no zeros)")
            else:
                # No data: linear 0-1 with potential watermark
                chart.y_axis.scaling = Scaling(logBase=None, min=0, orientation="minMax")
                if needs_watermark:
                    chart.y_axis.scaling.min = 1
                    chart.y_axis.scaling.max = 10
                print(f"     Axis: Linear scale 0-1 (no data, watermark={needs_watermark})")
            
            # Ensure axes are visible and formatted
            chart.y_axis.delete = False
            chart.y_axis.numFmt = 'General'  # Show numbers on Y-axis
            chart.x_axis.delete = False
            
            # X-axis formatting - dd-mmm format with tick marks
            chart.x_axis.number_format = 'dd-mmm'  # e.g., 06-Jun, 07-Jun
            chart.x_axis.tickLblPos = "low"  # Position labels below axis
            chart.x_axis.majorTickMark = "out"  # Tick marks outside
            
            # Configure gridlines: HORIZONTAL ONLY in chart area
            chart.y_axis.majorGridlines = None  # Remove default
            from openpyxl.chart.shapes import GraphicalProperties as ChartGraphicalProperties
            from openpyxl.drawing.line import LineProperties as ChartLineProperties
            
            # Add horizontal gridlines
            if chart.y_axis.majorGridlines is None:
                from openpyxl.chart.axis import ChartLines
                chart.y_axis.majorGridlines = ChartLines()
            
            # Style gridlines (light gray)
            gridline_props = ChartGraphicalProperties()
            gridline_line = ChartLineProperties()
            from openpyxl.drawing.fill import ColorChoice
            gridline_line.solidFill = ColorChoice(srgbClr="D9D9D9")  # Light gray
            gridline_line.w = 9525  # Line width
            gridline_props.line = gridline_line
            chart.y_axis.majorGridlines.spPr = gridline_props
            
            # Remove X-axis gridlines (vertical lines)
            chart.x_axis.majorGridlines = None
            chart.x_axis.minorGridlines = None
            
            # Chart size: 13cm width × 7.5cm height
            # Convert cm to Excel units (1 cm ≈ 36 points ≈ 0.393701 inches)
            # Excel chart width/height units are approximately in cm
            chart.width = 13  # 13 cm
            chart.height = 7.5  # 7.5 cm
            
            # Auto scale X-axis from data (from VBA)
            # Note: openpyxl auto-scales by default
            
            xvalues = Reference(ws, min_col=current_col, min_row=4, max_row=max_row)
            yvalues = Reference(ws, min_col=current_col + 1, min_row=4, max_row=max_row)
            series = Series(yvalues, xvalues, title_from_data=False)
            chart.series.append(series)
            
            if chart.series:
                s = chart.series[0]
                s.graphicalProperties = GraphicalProperties()
                s.graphicalProperties.line = LineProperties(w=19050)
                s.graphicalProperties.line.solidFill = ColorChoice(srgbClr="F4B042")
                s.graphicalProperties.shadow = None
                
                # Force straight lines (no smoothing) - from VBA
                s.smooth = False
                
                from openpyxl.chart.marker import Marker
                s.marker = Marker('circle')
                s.marker.size = 7
                s.marker.graphicalProperties = GraphicalProperties()
                s.marker.graphicalProperties.solidFill = ColorChoice(srgbClr="F4B042")
                s.marker.graphicalProperties.line = LineProperties()
                s.marker.graphicalProperties.line.solidFill = ColorChoice(srgbClr="F4B042")
            
            # Chart size already set above (13cm × 7.5cm)
            # Legend: None (removed as per image 4)
            chart.legend = None
            
            # Chart area formatting
            chart.graphicalProperties = GraphicalProperties()
            chart.graphicalProperties.shadow = None
            
            # Remove border (from VBA)
            # Note: openpyxl doesn't have direct border control like VBA, but we ensure no shadow
            
            # Dynamic chart positioning - 3 charts per row
            # Calculate row and column based on chart_count
            charts_per_row = 3
            row_num = chart_count // charts_per_row
            col_num = chart_count % charts_per_row
            
            # Calculate cell position for 13cm × 7.5cm charts
            # 13cm width ≈ 14 Excel columns
            # 7.5cm height ≈ 16 Excel rows
            
            # Use proper column letter conversion (A, O, AC for 3 columns)
            from openpyxl.utils import get_column_letter
            start_col_number = 1 + (col_num * 14)  # Column 1, 15, 29
            start_col_letter = get_column_letter(start_col_number)
            start_row = 5 + (row_num * 17)  # Row 5, 22, 39, etc.
            
            position = f"{start_col_letter}{start_row}"
            print(f"     → Placing chart at position {position} (chart #{chart_count + 1})")
            
            ws.add_chart(chart, position)
            
            # Add watermark if needed (from VBA)
            if needs_watermark:
                print(f"     → Adding 'no data available' watermark")
                self._add_watermark_to_chart(ws, position, site_name)
            
            sites_found.append(site_name)
            chart_count += 1
            current_col += 4
        
        # Summary logging
        print(f"\n{'='*60}")
        print(f"FIGURE 3.5 SUMMARY:")
        print(f"  ✅ Charts created: {chart_count}")
        print(f"  ✅ Sites included: {sites_found}")
        if sites_no_data:
            print(f"  ⚠️ Sites with no bat calls (empty charts): {sites_no_data}")
        if sites_missing:
            print(f"  ❌ Sites not found in processed_data: {sites_missing}")
        print(f"{'='*60}\n")
        
        return chart_count > 0
    
    def _add_watermark_to_chart(self, ws, chart_position: str, site_name: str):
        """
        Add a 'no data available' watermark to a chart
        Mimics VBA: diagonal gray text, semi-transparent, centered
        
        Note: openpyxl's textbox support is limited compared to VBA,
        so this creates a text cell overlay instead
        """
        try:
            # Parse chart position (e.g., "A5")
            import re
            match = re.match(r'([A-Z]+)(\d+)', chart_position)
            if not match:
                return
            
            col_letter = match.group(1)
            row_num = int(match.group(2))
            
            # Calculate approximate center of chart area
            # Charts are ~11 units wide × 8 units tall, roughly 6 columns × 15 rows
            center_col_offset = 3  # Middle of chart horizontally
            center_row_offset = 7  # Middle of chart vertically
            
            # Convert column letter to number
            from openpyxl.utils import column_index_from_string
            col_num = column_index_from_string(col_letter)
            
            watermark_col = col_num + center_col_offset
            watermark_row = row_num + center_row_offset
            
            # Add watermark text in cell
            cell = ws.cell(row=watermark_row, column=watermark_col)
            cell.value = "no data available"
            cell.font = Font(name='Montserrat', size=14, bold=True, color='B4B4B4')  # Gray
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Note: Excel rotation and transparency are limited in openpyxl
            # The VBA version has: rotation=315°, transparency=0.4
            # This is a simplified version that at least shows the text
            
            print(f"       Added watermark at row {watermark_row}, col {watermark_col}")
            
        except Exception as e:
            print(f"       ⚠️ Could not add watermark: {e}")
    
    def _add_table_sheet(self, table_data: dict):
        """Add a single table as a worksheet"""
        
        table_id = table_data['table_id']
        title = table_data['title']
        df = table_data['dataframe']
        notes = table_data.get('notes', [])
        
        # Create sheet name (max 31 chars)
        sheet_name = f"Table_{table_id}".replace('.', '-')[:31]
        ws = self.workbook.create_sheet(title=sheet_name)
        
        # Add title with Biologic branding
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=14, color=self.BIOLOGIC_PRIMARY)
        
        # Add Biologic footer
        ws['A2'] = "Biologic Environmental Survey"
        ws['A2'].font = Font(italic=True, size=9, color=self.BIOLOGIC_SECONDARY)
        
        # Add notes and determine starting row for data
        current_row = 3
        for note in notes:
            ws[f'A{current_row}'] = note
            ws[f'A{current_row}'].font = Font(italic=True, size=9, color=self.BIOLOGIC_SECONDARY)
            current_row += 1
        
        # Set start_row for data (after title and notes)
        start_row = current_row + 1
        
        # Add data
        if not df.empty:
            # Write headers with Biologic styling
            for col_idx, col_name in enumerate(df.columns, start=1):
                cell = ws.cell(row=start_row, column=col_idx, value=col_name)
                cell.font = Font(bold=True, color="FFFFFF")  # White text
                cell.fill = PatternFill(start_color=self.BIOLOGIC_PRIMARY, 
                                       end_color=self.BIOLOGIC_PRIMARY, fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Write data rows
            for row_idx, row_data in enumerate(df.itertuples(index=False), start=start_row + 1):
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    
                    # Special formatting for certain cells
                    if table_id == '3.8' and col_idx > 2:  # Ghost bat matrix
                        if value == 'R':
                            cell.fill = PatternFill(start_color="FFB6B6", end_color="FFB6B6", fill_type="solid")
                        elif value == 'C':
                            cell.fill = PatternFill(start_color="B6FFB6", end_color="B6FFB6", fill_type="solid")
                        elif value == 'NC':
                            cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            
            # Auto-adjust column widths
            for column_cells in ws.columns:
                length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)
        else:
            ws[f'A{start_row}'] = "No data available for this configuration"
            ws[f'A{start_row}'].font = Font(italic=True)
    
    def _add_audit_sheets(self, audit_reports: dict):
        """Add audit report sheets"""
        
        # Date coverage
        if 'date_coverage' in audit_reports and not audit_reports['date_coverage'].empty:
            ws = self.workbook.create_sheet(title="Audit_Date_Coverage")
            ws['A1'] = "Date Coverage Audit"
            ws['A1'].font = Font(bold=True, size=14)
            
            df = audit_reports['date_coverage']
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=3):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 3:  # Header row
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            
            # Auto-adjust widths
            for column in ws.columns:
                length = max(len(str(cell.value)) if cell.value else 0 for cell in column)
                ws.column_dimensions[column[0].column_letter].width = min(length + 2, 50)
        
        # Quality checks
        if 'quality_checks' in audit_reports and not audit_reports['quality_checks'].empty:
            ws = self.workbook.create_sheet(title="Audit_Quality_Checks")
            ws['A1'] = "Quality Control Checks"
            ws['A1'].font = Font(bold=True, size=14)
            
            df = audit_reports['quality_checks']
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=3):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 3:  # Header row
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                    
                    # Color code status
                    if value and '✅' in str(value):
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif value and '⚠️' in str(value):
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            
            # Auto-adjust widths
            for column in ws.columns:
                length = max(len(str(cell.value)) if cell.value else 0 for cell in column)
                ws.column_dimensions[column[0].column_letter].width = min(length + 2, 50)
        
        # Missing data
        if 'missing_data' in audit_reports and not audit_reports['missing_data'].empty:
            ws = self.workbook.create_sheet(title="Audit_Missing_Data")
            ws['A1'] = "Missing Data Summary"
            ws['A1'].font = Font(bold=True, size=14)
            
            df = audit_reports['missing_data']
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=3):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 3:  # Header row
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            
            # Auto-adjust widths
            for column in ws.columns:
                length = max(len(str(cell.value)) if cell.value else 0 for cell in column)
                ws.column_dimensions[column[0].column_letter].width = min(length + 2, 50)
    
    def _add_temperature_humidity_charts(self, processed_data: dict, config: dict) -> int:
        """
        Add temperature and humidity charts for continuous monitoring sites
        Creates 8 charts total: 4 temperature + 4 humidity (one per continuous site)
        
        Args:
            processed_data: Dictionary with processed site data
            config: Configuration with site selections
            
        Returns:
            Number of charts successfully created
        """
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
        from io import BytesIO
        
        # Configure matplotlib
        plt.rcParams['figure.dpi'] = 300
        plt.rcParams['savefig.dpi'] = 300
        plt.rcParams['font.family'] = 'sans-serif'
        plt.rcParams['font.sans-serif'] = ['Montserrat', 'DejaVu Sans', 'Arial']
        
        continuous_sites = config.get('continuous_sites', [])
        if not continuous_sites:
            return 0
        
        chart_count = 0
        
        # Create charts for each site
        for idx, site_name in enumerate(continuous_sites[:4], 1):
            site_key = f"{site_name}_continuous"
            
            if site_key not in processed_data:
                continue
                
            site_data = processed_data[site_key]
            df = site_data.get('raw_data')
            
            if df is None or df.empty:
                continue
            
            # Create temperature chart
            print(f"✓ Creating temperature chart {idx} for {site_name}")
            temp_sheet_name = f"TempChart_{site_name}"[:31]
            ws_temp = self.workbook.create_sheet(title=temp_sheet_name)
            
            temp_png = self._create_temperature_chart_png(df, site_name)
            if temp_png:
                from openpyxl.drawing.image import Image as XLImage
                img = XLImage(BytesIO(temp_png))
                # 25cm × 16cm at 96 DPI ≈ 945 × 605 pixels
                img.width = 945
                img.height = 605
                ws_temp.add_image(img, 'A1')
                chart_count += 1
                print(f"  ✅ Temperature chart created for {site_name}")
            
            # Create humidity chart
            print(f"✓ Creating humidity chart {idx} for {site_name}")
            hum_sheet_name = f"HumChart_{site_name}"[:31]
            ws_hum = self.workbook.create_sheet(title=hum_sheet_name)
            
            hum_png = self._create_humidity_chart_png(df, site_name)
            if hum_png:
                from openpyxl.drawing.image import Image as XLImage
                img = XLImage(BytesIO(hum_png))
                img.width = 945
                img.height = 605
                ws_hum.add_image(img, 'A1')
                chart_count += 1
                print(f"  ✅ Humidity chart created for {site_name}")
        
        return chart_count
    
    def _create_temperature_chart_png(self, df, site_name):
        """
        Create temperature chart matching the example style
        
        Args:
            df: DataFrame with raw site data
            site_name: Name of the site
            
        Returns:
            bytes: PNG image data, or None if failed
        """
        import matplotlib.pyplot as plt
        import matplotlib.dates as mdates
        import pandas as pd
        from io import BytesIO
        
        try:
            # Extract data
            date_col = df.columns[0]
            dates = pd.to_datetime(df[date_col], errors='coerce')
            
            # Column indices based on specification:
            # ambient_temp_min_c (Q) = column index 16
            # ambient_temp_max_c (R) = column index 17
            # roost_temperature_min (AN) = column index 39
            # roost_temperature_max (AO) = column index 40
            
            ambient_temp_min = pd.to_numeric(df.iloc[:, 16], errors='coerce') if len(df.columns) > 16 else pd.Series()
            ambient_temp_max = pd.to_numeric(df.iloc[:, 17], errors='coerce') if len(df.columns) > 17 else pd.Series()
            roost_temp_min = pd.to_numeric(df.iloc[:, 39], errors='coerce') if len(df.columns) > 39 else pd.Series()
            roost_temp_max = pd.to_numeric(df.iloc[:, 40], errors='coerce') if len(df.columns) > 40 else pd.Series()
            
            # Create figure: 25cm × 16cm
            # Convert to inches: 25cm ≈ 9.84", 16cm ≈ 6.30"
            fig, ax = plt.subplots(figsize=(9.84, 6.30))
            fig.patch.set_alpha(0)  # Transparent background
            ax.patch.set_alpha(0)
            
            # Define colors matching example images
            AMBIENT_COLOR = '#D4AF37'  # Gold/yellow
            ROOST_COLOR = '#2F4F4F'    # Dark teal
            
            # Plot lines
            if not ambient_temp_max.empty and ambient_temp_max.notna().any():
                ax.plot(dates, ambient_temp_max, color=AMBIENT_COLOR, linewidth=2.0, 
                       linestyle='-', label='Ambient Temp (Max) (C°)', zorder=3)
            
            if not ambient_temp_min.empty and ambient_temp_min.notna().any():
                ax.plot(dates, ambient_temp_min, color=AMBIENT_COLOR, linewidth=2.0,
                       linestyle='--', label='Ambient Temp (Min) (C°)', zorder=3)
            
            if not roost_temp_max.empty and roost_temp_max.notna().any():
                ax.plot(dates, roost_temp_max, color=ROOST_COLOR, linewidth=2.0,
                       linestyle='-', label='Roost Temp Max (C°)', zorder=2)
            
            if not roost_temp_min.empty and roost_temp_min.notna().any():
                ax.plot(dates, roost_temp_min, color=ROOST_COLOR, linewidth=2.0,
                       linestyle='--', label='Roost Temp Min (C°)', zorder=2)
            
            # Configure axes
            ax.set_ylabel('Temperature °C', fontsize=11, fontname='Montserrat')
            ax.set_xlabel('')
            
            # Horizontal gridlines only
            ax.grid(True, axis='y', linestyle='--', linewidth=0.5, alpha=0.5, color='gray')
            ax.grid(False, axis='x')
            
            # Format x-axis as mm/yyyy
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%m/%Y'))
            ax.xaxis.set_major_locator(mdates.MonthLocator(interval=1))
            plt.setp(ax.xaxis.get_majorticklabels(), rotation=0, ha='center', fontsize=10)
            
            # Set x-axis limits
            if not dates.empty:
                date_min = dates.min()
                date_max = dates.max()
                if pd.notna(date_min) and pd.notna(date_max):
                    ax.set_xlim(date_min, date_max)
            
            # Y-axis auto-scale
            ax.tick_params(axis='y', labelsize=10)
            
            # Add title
            ax.set_title(site_name, fontsize=12, fontname='Montserrat', pad=15)
            
            # Legend below chart
            handles, labels = ax.get_legend_handles_labels()
            ax.legend(handles, labels, loc='upper center', bbox_to_anchor=(0.5, -0.08),
                     ncol=4, frameon=False, fontsize=10)
            
            # Tight layout
            plt.tight_layout()
            plt.subplots_adjust(bottom=0.15)
            
            # Save to bytes
            buf = BytesIO()
            plt.savefig(buf, format='png', dpi=300, transparent=True, bbox_inches='tight')
            plt.close(fig)
            buf.seek(0)
            
            return buf.getvalue()
            
        except Exception as e:
            print(f"  ❌ Error creating temperature chart for {site_name}: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def _create_humidity_chart_png(self, df, site_name):
        """
        Create humidity chart matching the example style
        
        Args:
            df: DataFrame with raw site data
            site_name: Name of the site
            
        Returns:
            bytes: PNG image data, or None if failed
        """
        import matplotlib.pyplot as plt
        import matplotlib.dates as mdates
        import pandas as pd
        from io import BytesIO
        
        try:
            # Extract data
            date_col = df.columns[0]
            dates = pd.to_datetime(df[date_col], errors='coerce')
            
            # Column indices based on specification:
            # ambient_9am_rh (T) = column index 19 -> Ambient RH Max
            # ambient_3pm_rh (U) = column index 20 -> Ambient RH Min
            # roost_rh_max (AQ) = column index 42
            # roost_rh_min (AP) = column index 41
            
            ambient_rh_max = pd.to_numeric(df.iloc[:, 19], errors='coerce') if len(df.columns) > 19 else pd.Series()
            ambient_rh_min = pd.to_numeric(df.iloc[:, 20], errors='coerce') if len(df.columns) > 20 else pd.Series()
            roost_rh_max = pd.to_numeric(df.iloc[:, 42], errors='coerce') if len(df.columns) > 42 else pd.Series()
            roost_rh_min = pd.to_numeric(df.iloc[:, 41], errors='coerce') if len(df.columns) > 41 else pd.Series()
            
            # Create figure: 25cm × 16cm
            fig, ax = plt.subplots(figsize=(9.84, 6.30))
            fig.patch.set_alpha(0)  # Transparent background
            ax.patch.set_alpha(0)
            
            # Define colors matching example images
            AMBIENT_COLOR = '#D4AF37'  # Gold/yellow
            ROOST_COLOR = '#2F4F4F'    # Dark teal
            
            # Plot lines
            if not ambient_rh_max.empty and ambient_rh_max.notna().any():
                ax.plot(dates, ambient_rh_max, color=AMBIENT_COLOR, linewidth=2.0,
                       linestyle='-', label='Ambient RH Max', zorder=3)
            
            if not ambient_rh_min.empty and ambient_rh_min.notna().any():
                ax.plot(dates, ambient_rh_min, color=AMBIENT_COLOR, linewidth=2.0,
                       linestyle='--', label='Ambient RH Min', zorder=3)
            
            if not roost_rh_max.empty and roost_rh_max.notna().any():
                ax.plot(dates, roost_rh_max, color=ROOST_COLOR, linewidth=2.0,
                       linestyle='-', label='Roost RH Max', zorder=2)
            
            if not roost_rh_min.empty and roost_rh_min.notna().any():
                ax.plot(dates, roost_rh_min, color=ROOST_COLOR, linewidth=2.0,
                       linestyle='--', label='Roost RH Min', zorder=2)
            
            # Configure axes
            ax.set_ylabel('Humidity %', fontsize=11, fontname='Montserrat')
            ax.set_xlabel('')
            
            # Horizontal gridlines only
            ax.grid(True, axis='y', linestyle='--', linewidth=0.5, alpha=0.5, color='gray')
            ax.grid(False, axis='x')
            
            # Format x-axis as mm/yyyy
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%m/%Y'))
            ax.xaxis.set_major_locator(mdates.MonthLocator(interval=1))
            plt.setp(ax.xaxis.get_majorticklabels(), rotation=0, ha='center', fontsize=10)
            
            # Set x-axis limits
            if not dates.empty:
                date_min = dates.min()
                date_max = dates.max()
                if pd.notna(date_min) and pd.notna(date_max):
                    ax.set_xlim(date_min, date_max)
            
            # Y-axis auto-scale
            ax.tick_params(axis='y', labelsize=10)
            
            # Add title
            ax.set_title(site_name, fontsize=12, fontname='Montserrat', pad=15)
            
            # Legend below chart
            handles, labels = ax.get_legend_handles_labels()
            ax.legend(handles, labels, loc='upper center', bbox_to_anchor=(0.5, -0.08),
                     ncol=4, frameon=False, fontsize=10)
            
            # Tight layout
            plt.tight_layout()
            plt.subplots_adjust(bottom=0.15)
            
            # Save to bytes
            buf = BytesIO()
            plt.savefig(buf, format='png', dpi=300, transparent=True, bbox_inches='tight')
            plt.close(fig)
            buf.seek(0)
            
            return buf.getvalue()
            
        except Exception as e:
            print(f"  ❌ Error creating humidity chart for {site_name}: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def export_table_to_csv(self, table_data: dict) -> str:
        """Export a single table to CSV format"""
        df = table_data['dataframe']
        return df.to_csv(index=False)
    
    # ========================================================================
    # BAT ACTIVITY STATEMENTS METHODS
    # ========================================================================
    
    def _add_bat_activity_statements(self, master_data_df: pd.DataFrame):
        """
        Add Bat Activity Statements sheet with site_id, date, species, and formatted statement (4 columns)
        
        Args:
            master_data_df: DataFrame with columns including:
                - date
                - gb_first_call_time, gb_last_call_time, gb_total_calls, gb_roosting_indicated
                - plnb_first_call_time, plnb_last_call_time, plnb_total_calls, plnb_roosting_indicated
                - site_id (optional - if not present, will use row index)
        """
        if master_data_df is None or master_data_df.empty:
            print("⚠️ No master data provided for bat activity statements")
            return
        
        print("Creating Bat Activity Statements sheet...")
        
        # Create output list
        statements = []
        
        # Determine if site_id column exists
        has_site_id = 'site_id' in master_data_df.columns
        
        # Process each row for Ghost Bat (GB) and PLNB
        for idx, row in master_data_df.iterrows():
            raw_site_id = row.get('site_id', f'Site_{idx+1}') if has_site_id else f'Row_{idx+1}'
            
            # Clean site name - just use as-is for McPhee (CMPC-08, etc.)
            site_id = raw_site_id
            
            # Format date as dd/mm/yyyy
            date_raw = row.get('date', '')
            date = self._format_date_ddmmyyyy(date_raw)
            
            # Check if row has GB data (any GB column has a value)
            has_gb_data = any(pd.notna(row.get(col)) and str(row.get(col)).strip() != '' 
                            for col in ['gb_first_call_time', 'gb_last_call_time', 'gb_total_calls', 'gb_roosting_indicated'])
            
            if has_gb_data:
                statement = self._create_gb_statement(row)
                statements.append({
                    'site_id': site_id,
                    'date': date,
                    'species': 'GB',
                    'statement': statement
                })
            
            # Check if row has PLNB data
            has_plnb_data = any(pd.notna(row.get(col)) and str(row.get(col)).strip() != ''
                              for col in ['plnb_total_calls', 'plnb_roosting_indicated'])
            
            if has_plnb_data:
                statement = self._create_plnb_statement(row)
                statements.append({
                    'site_id': site_id,
                    'date': date,
                    'species': 'PLNB',
                    'statement': statement
                })
        
        # Create DataFrame with 4 columns
        statements_df = pd.DataFrame(statements)
        
        # Add to workbook
        ws = self.workbook.create_sheet(title="Bat Activity Statements")
        
        # Write headers - 4 COLUMNS
        headers = ['site_id', 'date', 'species', 'statement']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(name='Montserrat', size=11, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.BIOLOGIC_PRIMARY, 
                                   end_color=self.BIOLOGIC_PRIMARY, fill_type='solid')
            cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Write data
        for row_idx, row_data in enumerate(statements_df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = Font(name='Montserrat', size=10)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                # Alternate row colors
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color=self.BIOLOGIC_BACKGROUND,
                                          end_color=self.BIOLOGIC_BACKGROUND, fill_type='solid')
        
        # Set column widths - 4 COLUMNS
        ws.column_dimensions['A'].width = 15  # site_id
        ws.column_dimensions['B'].width = 12  # date
        ws.column_dimensions['C'].width = 10  # species
        ws.column_dimensions['D'].width = 85  # statement
        
        # Set row height for header
        ws.row_dimensions[1].height = 20
        
        print(f"✅ Added Bat Activity Statements sheet with {len(statements)} records")
    
    def _format_date_ddmmyyyy(self, date_value) -> str:
        """
        Format date as dd/mm/yyyy
        """
        if pd.isna(date_value) or date_value == '':
            return ''
        
        try:
            if isinstance(date_value, str):
                # Parse string date
                dt = pd.to_datetime(date_value)
                return dt.strftime('%d/%m/%Y')
            else:
                # Already a datetime
                return pd.to_datetime(date_value).strftime('%d/%m/%Y')
        except:
            return str(date_value)
    
    def _format_time_hmm(self, time_value) -> str:
        """
        Format time as h:mm (not zero-padded hour)
        Examples: 9:05, 14:30
        Handles input like "5:46", "05:46", "5'46" and converts to proper format
        """
        if pd.isna(time_value) or time_value == '':
            return "n/a"
        
        try:
            # Convert to string first
            time_str = str(time_value).strip()
            
            # If it's already a time string like "5:46" or "05:46" or "5'46"
            if ':' in time_str or "'" in time_str or '\u2018' in time_str or '\u2019' in time_str:
                # Replace any apostrophe-like characters with colon
                time_str = time_str.replace("'", ':').replace('\u2018', ':').replace('\u2019', ':')
                
                # Split on colon
                parts = time_str.split(':')
                if len(parts) == 2:
                    try:
                        hour = int(parts[0])
                        minute = int(parts[1])
                        return f"{hour}:{minute:02d}"
                    except ValueError:
                        pass
            
            # Try parsing as datetime
            dt = pd.to_datetime(time_value)
            hour = dt.hour
            minute = dt.minute
            return f"{hour}:{minute:02d}"
            
        except:
            # If all else fails, return cleaned string
            result = str(time_value).replace("'", ':').replace('\u2018', ':').replace('\u2019', ':')
            return result if result else "n/a"
    
    def _format_total_calls(self, calls_value) -> str:
        """
        Format total calls as whole number (integer)
        """
        if pd.isna(calls_value) or calls_value == '':
            return "n/a"
        
        calls_str = str(calls_value).strip()
        
        # Handle special cases like ">15"
        if calls_str.startswith('>'):
            return calls_str
        
        try:
            # Convert to integer (removes decimals)
            return str(int(float(calls_str)))
        except:
            return calls_str
    
    def _create_gb_statement(self, row) -> str:
        """Create Ghost Bat activity statement from row data"""
        # Extract and format First Call time using h:mm format
        first_call = row.get('gb_first_call_time')
        first_call_str = self._format_time_hmm(first_call)
        
        # Extract and format Last Call time using h:mm format
        last_call = row.get('gb_last_call_time')
        last_call_str = self._format_time_hmm(last_call)
        
        # Extract Total Calls as whole number
        total_calls = row.get('gb_total_calls')
        total_calls_str = self._format_total_calls(total_calls)
        
        # Extract Roosting status
        roosting = row.get('gb_roosting_indicated', '')
        roosting_str = self._format_roosting_status(roosting)
        
        # Combine into statement
        statement = f"First call {first_call_str} ; Last call {last_call_str} ; Total calls {total_calls_str} ; {roosting_str}"
        
        return statement
    
    def _create_plnb_statement(self, row) -> str:
        """Create PLNB activity statement from row data"""
        # PLNB always has n/a for first and last call times
        first_call_str = "n/a"
        last_call_str = "n/a"
        
        # Extract Total Calls as whole number
        total_calls = row.get('plnb_total_calls')
        total_calls_str = self._format_total_calls(total_calls)
        
        # Extract Roosting status
        roosting = row.get('plnb_roosting_indicated', '')
        roosting_str = self._format_roosting_status(roosting)
        
        # Combine into statement
        statement = f"First call {first_call_str} ; Last call {last_call_str} ; Total calls {total_calls_str} ; {roosting_str}"
        
        return statement
    
    def _format_roosting_status(self, roosting_value) -> str:
        """
        Convert roosting indicator to statement text
        
        Args:
            roosting_value: Can be 1, 0, "YES", "NO", or empty
            
        Returns:
            Formatted roosting statement
        """
        if pd.isna(roosting_value) or roosting_value == '':
            return "Timing of calls not indicative of roosting"
        
        # Convert to string and check
        roosting_str = str(roosting_value).strip().upper()
        
        if roosting_str in ['1', 'YES', 'TRUE']:
            return "Timing of calls indicative of roosting"
        elif roosting_str in ['0', 'NO', 'FALSE']:
            return "Timing of calls not indicative of roosting"
        else:
            return "Timing of calls not indicative of roosting"
